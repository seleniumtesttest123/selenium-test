import openpyxl

# file="C:\\Users\\vinay\\OneDrive\\Documents\\hi.xlsx"
# workbook=openpyxl.load_workbook(file)
# sheet=workbook.active #or sheet=worksheet("sheet")
#
# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(r,c).value="welcome"
# workbook.save(file)

file = "C:\\Users\\vinay\\OneDrive\\Documents\\hi.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook.active  # or sheet=worksheet("sheet")

sheet.cell(1, 1).value = 123
sheet.cell(1, 2).value = "smith"
sheet.cell(1, 3).value = "Engineer"

sheet.cell(2, 1).value = 456
sheet.cell(2, 2).value = "jorge"
sheet.cell(2, 3).value = "Manager"

sheet.cell(3, 1).value = 789
sheet.cell(3, 2).value = "danny"
sheet.cell(3, 3).value = "Developer"

workbook.save(file)
